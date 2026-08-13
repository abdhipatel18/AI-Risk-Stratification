import os
import logging
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

class ExcelReportGenerator:
    def __init__(self, output_dir="data/output"):
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)

    def generate_excel_report(self, df_processed, df_quality_report, filename="patient_risk_prioritization_report.xlsx"):
        """
        Generates a multi-tab formatted Excel workbook complying with PDF specifications.
        """
        output_path = os.path.join(self.output_dir, filename)
        wb = openpyxl.Workbook()
        wb.remove(wb.active) # Remove default sheet

        # 1. Executive Summary Sheet
        ws_exec = wb.create_sheet(title="Executive Summary")
        self._build_executive_summary(ws_exec, df_processed, df_quality_report)

        # 2. Patient Priority List Sheet
        ws_priority = wb.create_sheet(title="Patient Priority List")
        self._write_styled_dataframe(ws_priority, df_processed)

        # 3. One-Month Scheduling List Sheet
        ws_1m = wb.create_sheet(title="One-Month Scheduling List")
        df_1m = df_processed[df_processed["recommended_scheduling_window"] == "Schedule Within One Month"]
        self._write_styled_dataframe(ws_1m, df_1m)

        # 4. Two-Month Scheduling List Sheet
        ws_2m = wb.create_sheet(title="Two-Month Scheduling List")
        df_2m = df_processed[df_processed["recommended_scheduling_window"] == "Schedule Within Two Months"]
        self._write_styled_dataframe(ws_2m, df_2m)

        # 5. Three-Month Scheduling List Sheet
        ws_3m = wb.create_sheet(title="Three-Month Scheduling List")
        df_3m = df_processed[df_processed["recommended_scheduling_window"] == "Schedule Within Three Months"]
        self._write_styled_dataframe(ws_3m, df_3m)

        # 6. Care-Gap Summary Sheet
        ws_gap = wb.create_sheet(title="Care-Gap Summary")
        df_gap_summary = self._build_care_gap_summary(df_processed)
        self._write_styled_dataframe(ws_gap, df_gap_summary)

        # 7. Data-Quality Report Sheet
        ws_dq = wb.create_sheet(title="Data-Quality Report")
        self._write_styled_dataframe(ws_dq, df_quality_report)

        wb.save(output_path)
        logger.info(f"Successfully generated formatted Excel report at '{output_path}'")
        return output_path

    def _build_executive_summary(self, ws, df, df_dq):
        ws.title = "Executive Summary"
        
        # Styles
        title_font = Font(name="Calibri", size=16, bold=True, color="1F4E78")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        bold_font = Font(name="Calibri", size=11, bold=True)
        center_align = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("A1:D1")
        ws["A1"] = "Executive Summary: Patient Risk Stratification & Scheduling Prioritization"
        ws["A1"].font = title_font

        total_patients = len(df)
        high_risk = len(df[df["risk_category"] == "High Risk"])
        med_risk = len(df[df["risk_category"] == "Medium Risk"])
        low_risk = len(df[df["risk_category"] == "Lower Risk"])
        scheduled_no_act = len(df[df["risk_category"] == "No Immediate Action or Already Scheduled"])

        sched_1m = len(df[df["recommended_scheduling_window"] == "Schedule Within One Month"])
        sched_2m = len(df[df["recommended_scheduling_window"] == "Schedule Within Two Months"])
        sched_3m = len(df[df["recommended_scheduling_window"] == "Schedule Within Three Months"])
        already_sched = len(df[df["recommended_scheduling_window"] == "Already Scheduled or No Scheduling Needed"])
        dq_issues = len(df_dq)

        summary_data = [
            ["Metric Category", "Metric Name", "Count", "Percentage / Details"],
            ["Overview", "Total Patients Analyzed", total_patients, "100.0%"],
            ["Risk Stratification", "High Risk Patients", high_risk, f"{high_risk/max(1,total_patients):.1%}"],
            ["Risk Stratification", "Medium Risk Patients", med_risk, f"{med_risk/max(1,total_patients):.1%}"],
            ["Risk Stratification", "Lower Risk Patients", low_risk, f"{low_risk/max(1,total_patients):.1%}"],
            ["Risk Stratification", "No Immediate Action / Scheduled", scheduled_no_act, f"{scheduled_no_act/max(1,total_patients):.1%}"],
            ["Scheduling Recommendation", "Schedule Within 1 Month", sched_1m, f"{sched_1m/max(1,total_patients):.1%}"],
            ["Scheduling Recommendation", "Schedule Within 2 Months", sched_2m, f"{sched_2m/max(1,total_patients):.1%}"],
            ["Scheduling Recommendation", "Schedule Within 3 Months", sched_3m, f"{sched_3m/max(1,total_patients):.1%}"],
            ["Scheduling Recommendation", "Already Scheduled / No Need", already_sched, f"{already_sched/max(1,total_patients):.1%}"],
            ["Data Quality", "Missing / Incomplete Data Records", dq_issues, f"{dq_issues} quality flags logged"]
        ]

        for r_idx, row in enumerate(summary_data, start=3):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == 3:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                elif c_idx == 1:
                    cell.font = bold_font

        self._auto_fit_columns(ws)

    def _build_care_gap_summary(self, df):
        summary_rows = []
        # Count highest priority gaps & overdue gaps
        gap_counts = df["highest_priority_care_gap"].value_counts()
        for gap_name, count in gap_counts.items():
            if gap_name == "None":
                continue
            sub_df = df[df["highest_priority_care_gap"] == gap_name]
            avg_weight = round(sub_df["max_individual_care_gap_weight"].mean(), 1)
            overdue_count = sub_df["num_overdue_care_gaps"].sum()
            in_1m = len(sub_df[sub_df["recommended_scheduling_window"] == "Schedule Within One Month"])

            summary_rows.append({
                "care_gap_type": gap_name,
                "number_of_affected_patients": count,
                "average_care_gap_weight": avg_weight,
                "number_overdue": overdue_count,
                "number_in_one_month_group": in_1m
            })

        return pd.DataFrame(summary_rows) if summary_rows else pd.DataFrame(columns=["care_gap_type", "number_of_affected_patients", "average_care_gap_weight", "number_overdue", "number_in_one_month_group"])

    def _write_styled_dataframe(self, ws, df):
        if df.empty:
            ws.append(["No records available."])
            return

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        thin_border = Border(left=Side(style='thin', color='D9D9D9'),
                             right=Side(style='thin', color='D9D9D9'),
                             top=Side(style='thin', color='D9D9D9'),
                             bottom=Side(style='thin', color='D9D9D9'))

        # Write Headers
        headers = list(df.columns)
        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Write Data
        for row in df.itertuples(index=False):
            ws.append(list(row))

        # Borders & formatting
        for row in ws.iter_rows(min_row=2, max_row=len(df)+1, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border

        self._auto_fit_columns(ws)

    def _auto_fit_columns(self, ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or "")
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)
